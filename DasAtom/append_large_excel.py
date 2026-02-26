import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ============ 辅助：从单个电路 xlsx 提取指标 ============
def extract_from_xlsx(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    d = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] and isinstance(row[0], str):
            d[row[0].strip()] = row[1]
    wb.close()
    return d

# ============ 读取大电路 dual 数据 ============
dual_dir = r'e:\coding\DasAtom_reading\DasAtom\res\dual_large\Rb2Re4'
base_dir = r'e:\coding\DasAtom_reading\DasAtom_Origin\res\baseline_large\Rb2Re4'
circs = ['qft_24', 'qv_24', 'random_24', 'star_24', 'qft_30']
g_counts = {'qft_24':588, 'qv_24':288, 'random_24':174, 'star_24':276, 'qft_30':915}
q_counts = {'qft_24':24, 'qv_24':24, 'random_24':24, 'star_24':24, 'qft_30':30}

large_data = []
for c in circs:
    f = os.path.join(dual_dir, f'{c}.qasm_rb2.xlsx')
    if os.path.exists(f):
        m = extract_from_xlsx(f)
        dist = m.get('Total distance', m.get('Total_distance', 0))
        fid = m.get('Fidelity', 0)
        embed_t = m.get('Embedding computation time', 0)
        mcts_t = m.get('MCTS search time', 0)
        tt = float(embed_t or 0) + float(mcts_t or 0)
        
        # Check if baseline completed
        fb = os.path.join(base_dir, f'{c}.qasm_rb2.xlsx')
        base_status = "超时 (>600s)"
        base_dist = None
        if os.path.exists(fb):
            try:
                bm = extract_from_xlsx(fb)
                if 'Total distance' in bm or 'Total_distance' in bm:
                    base_status = "完成 (但报告未生成)"
                    base_dist = bm.get('Total distance', bm.get('Total_distance', 0))
            except:
                pass
                
        large_data.append((c, q_counts[c], g_counts[c], base_status, dist, fid, tt))

# ============ 读取并更新现有的 benchmark_report.xlsx ============
out_file = r'e:\coding\DasAtom_reading\DasAtom\benchmark_report.xlsx'
wb = load_workbook(out_file)
ws = wb.active

# Skip to end
r_idx = ws.max_row + 2

sec_font = Font(name='微软雅黑', bold=True, size=10, color='2F5496')
sec_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
center = Alignment(horizontal='center', vertical='center')

ws.cell(row=r_idx, column=1, value='▼ 大电路极限测试 (24~30 比特，原版设 10 分钟超时)').font = sec_font
for col in range(1, 13):
    ws.cell(row=r_idx, column=col).fill = sec_fill
    ws.cell(row=r_idx, column=col).border = thin
r_idx += 1

data_f = Font(name='微软雅黑', size=10)
red_f = Font(name='微软雅黑', size=10, color='FF0000', italic=True)
green_f = Font(name='微软雅黑', size=10, color='008000', bold=True)

for circ, q, g, b_stat, d_dist, d_fid, d_t in large_data:
    ws.cell(row=r_idx, column=1, value=circ).font = data_f
    ws.cell(row=r_idx, column=2, value=q).font = data_f
    ws.cell(row=r_idx, column=3, value=f'门数:{g}').font = data_f
    
    ws.cell(row=r_idx, column=4, value=b_stat).font = red_f
    ws.cell(row=r_idx, column=5, value=d_dist).font = data_f
    ws.cell(row=r_idx, column=6, value='VF2 失败').font = red_f
    
    ws.cell(row=r_idx, column=7, value=b_stat).font = red_f
    ws.cell(row=r_idx, column=8, value=d_fid).font = data_f
    ws.cell(row=r_idx, column=8).number_format = '0.000000'
    ws.cell(row=r_idx, column=9, value='VF2 失败').font = red_f
    
    ws.cell(row=r_idx, column=10, value='>600.0s').font = red_f
    ws.cell(row=r_idx, column=11, value=d_t).font = data_f
    ws.cell(row=r_idx, column=11).number_format = '0.00'
    ws.cell(row=r_idx, column=12, value='∞ 倍').font = green_f
    
    for col in range(1, 13):
        ws.cell(row=r_idx, column=col).alignment = center
        ws.cell(row=r_idx, column=col).border = thin
    r_idx += 1

wb.save(out_file)
print("大电路追加完成！")
