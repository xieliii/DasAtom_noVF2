"""
对比 noVF2 引擎和已有基准(Baseline/Dual)的性能。
读取 medium_benchmark.xlsx (Baseline/Dual) 和 noVF2 summary.xlsx,
输出到新 Excel 和控制台表格。
"""
import openpyxl
import os
import sys

def load_noVF2_results(path):
    """从 noVF2 summary xlsx 读取结果"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    results = {}
    headers = [c.value for c in ws[1]]
    
    # 找到关键列的索引
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[h] = i
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        vals = [c.value for c in row]
        name = vals[0]
        if name is None or name == '' or not isinstance(name, str):
            continue
        # 去掉 .qasm 后缀
        name = name.replace('.qasm', '')
        
        results[name] = {
            'qubits': vals[col_map.get('Num Qubits', 1)],
            'cz_gates': vals[col_map.get('Num CZ Gates', 2)],
            'fidelity': vals[col_map.get('Fidelity', 4)],
            'move_distance': vals[col_map.get('Total Move Distance', 9)],
            'partitions': vals[col_map.get('Num Partitions', 11)],
            'time': vals[col_map.get('Elapsed Time (s)', 12)],
        }
    return results


def load_medium_benchmark(path):
    """从 medium_benchmark.xlsx 读取 Baseline 和 Dual 结果"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    
    results = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        vals = [c.value for c in row]
        name = vals[0]
        if name is None or not isinstance(name, str):
            continue
        
        results[name] = {
            'qubits': vals[1],
            'cz_gates': vals[2],
            'baseline_distance': vals[3],
            'dual_distance': vals[4],
            'dual_distance_reduction': vals[5],
            'baseline_fidelity': vals[6],
            'dual_fidelity': vals[7],
            'baseline_time': vals[8],
            'dual_time': vals[9],
        }
    return results


def main():
    novf2_path = r'res\noVF2_run3\Rb2Re4\medium_noVF2_summary.xlsx'
    medium_path = r'medium_benchmark.xlsx'
    
    if not os.path.exists(novf2_path):
        print(f"找不到 noVF2 结果: {novf2_path}")
        sys.exit(1)
    if not os.path.exists(medium_path):
        print(f"找不到基准文件: {medium_path}")
        sys.exit(1)
    
    novf2 = load_noVF2_results(novf2_path)
    medium = load_medium_benchmark(medium_path)
    
    print(f"\n{'='*100}")
    print(f"noVF2 vs Baseline/Dual 性能对比")
    print(f"{'='*100}")
    
    # 表头
    header = f"{'Circuit':<15} {'Q':>3} {'CZ':>4} | {'Baseline':>10} {'Dual':>10} {'noVF2':>10} | {'Baseline':>10} {'Dual':>10} {'noVF2':>10} | {'Baseline':>10} {'Dual':>10} {'noVF2':>10} {'加速比':>8}"
    print(f"\n{'':>24}      | {'--- 移动距离 ---':^32} | {'--- 保真度 ---':^32} | {'--- 运行时间(s) ---':^44}")
    print(header)
    print('-' * len(header))
    
    # 创建输出 Excel
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Comparison"
    out_ws.append([
        '电路', 'Qubits', 'CZ门数',
        'Baseline距离', 'Dual距离', 'noVF2距离', '距离变化(%)',
        'Baseline保真度', 'Dual保真度', 'noVF2保真度', '保真度变化(%)',
        'Baseline时间(s)', 'Dual时间(s)', 'noVF2时间(s)', '加速比(Dual/noVF2)'
    ])
    
    # 按 medium benchmark 中的电路顺序输出
    total_baseline_time = 0
    total_dual_time = 0
    total_novf2_time = 0
    
    for name, med in medium.items():
        if name not in novf2:
            continue
        nv = novf2[name]
        
        bl_dist = med.get('baseline_distance', '-')
        du_dist = med.get('dual_distance', '-')
        nv_dist = nv.get('move_distance', '-')
        
        bl_fid = med.get('baseline_fidelity', '-')
        du_fid = med.get('dual_fidelity', '-')
        nv_fid = nv.get('fidelity', '-')
        
        bl_time = med.get('baseline_time', '-')
        du_time = med.get('dual_time', '-')
        nv_time = nv.get('time', '-')
        
        # 加速比
        speedup = '-'
        if isinstance(du_time, (int, float)) and isinstance(nv_time, (int, float)) and nv_time > 0:
            speedup = f"{du_time / nv_time:.1f}x"
        
        # 距离变化
        dist_change = '-'
        if isinstance(du_dist, (int, float)) and isinstance(nv_dist, (int, float)) and du_dist > 0:
            dist_change = f"{(nv_dist - du_dist) / du_dist * 100:+.1f}%"
            
        # 保真度变化
        fid_change = '-'
        if isinstance(du_fid, (int, float)) and isinstance(nv_fid, (int, float)) and du_fid > 0:
            fid_change = f"{(nv_fid - du_fid) / du_fid * 100:+.1f}%"
        
        # 格式化输出
        def fmt(v, w=10):
            if isinstance(v, float):
                if abs(v) < 0.01:
                    return f"{v:>{w}.4f}"
                return f"{v:>{w}.3f}"
            return f"{str(v):>{w}}"
        
        print(f"{name:<15} {med['qubits']:>3} {med['cz_gates']:>4} | {fmt(bl_dist)} {fmt(du_dist)} {fmt(nv_dist)} | {fmt(bl_fid)} {fmt(du_fid)} {fmt(nv_fid)} | {fmt(bl_time)} {fmt(du_time)} {fmt(nv_time)} {speedup:>8}")
        
        # Excel 行
        out_ws.append([
            name, med['qubits'], med['cz_gates'],
            bl_dist, du_dist, nv_dist, dist_change,
            bl_fid, du_fid, nv_fid, fid_change,
            bl_time, du_time, nv_time, speedup
        ])
        
        if isinstance(bl_time, (int, float)):
            total_baseline_time += bl_time
        if isinstance(du_time, (int, float)):
            total_dual_time += du_time
        if isinstance(nv_time, (int, float)):
            total_novf2_time += nv_time
    
    print('-' * len(header))
    total_speedup = f"{total_dual_time / total_novf2_time:.1f}x" if total_novf2_time > 0 else '-'
    print(f"{'TOTAL':>24}      | {'':>32} | {'':>32} | {total_baseline_time:>10.3f} {total_dual_time:>10.3f} {total_novf2_time:>10.3f} {total_speedup:>8}")
    
    out_path = 'noVF2_vs_dual_benchmark.xlsx'
    out_wb.save(out_path)
    print(f"\n结果已保存到: {out_path}")


if __name__ == '__main__':
    main()
