import os
import math
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ============ 物理参数 (From DasAtom_fun.py) ============
T_cz = 0.2
T_eff = 1.5e6
T_trans = 20
Move_speed = 0.55
F_cz_val = 0.995
F_trans_val = 1.0  # 原始代码中默认设定光镊无错误 (1.0)

def extract_from_xlsx(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    d = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] and isinstance(row[0], str):
            d[row[0].strip()] = row[1]
    wb.close()
    return d

base_dir = r'e:\coding\DasAtom_reading\DasAtom_Origin\res\baseline_benchmark\Rb2Re4'
dual_dir = r'e:\coding\DasAtom_reading\DasAtom\res\dual_benchmark\Rb2Re4'

# 我们从主报告对应的25个电路来分析
circuits = [
    'linear_6', 'qft_6', 'qv_6', 'random_6', 'star_6',
    'linear_8', 'qft_8', 'qv_8', 'random_8', 'star_8',
    'linear_12', 'qft_12', 'qv_12', 'random_12', 'star_12',
    'linear_16', 'qft_16', 'qv_16', 'random_16', 'star_16',
    'linear_20', 'qft_20', 'qv_20', 'random_20', 'star_20'
]

data_rows = []

for circ in circuits:
    # Baseline
    fb = os.path.join(base_dir, f'{circ}.qasm_rb2.xlsx')
    base_data = extract_from_xlsx(fb) if os.path.exists(fb) else {}
    
    # Dual
    fd = os.path.join(dual_dir, f'{circ}.qasm_rb2.xlsx')
    dual_data = extract_from_xlsx(fd) if os.path.exists(fd) else {}
    
    for engine, md in [('Baseline (原版VF2)', base_data), ('Dual (纯血新版)', dual_data)]:
        if not md: continue
        
        num_q = md.get('Number of qubits (num_q)', md.get('Num qubits', 0))
        num_cz = md.get('Number of CZ gates', 0)
        
        idle = md.get('Idle time', 0)
        t_tot = md.get('Total_T (from fidelity calc)', md.get('Elapsed Time (s)', 0))
        if idle == 0 and 'Total_T (from fidelity calc)' in md:
            idle = num_q * t_tot - num_cz * T_cz
            
        # 根据原始代码 `num_trans += 4` 每个平行步
        # Number of moves 并不是 parallel steps，但我们直接根据抓到的 t_move / Move_speed 也能推。
        # 最稳妥的是：原代码 F_trans_val = 1.0，所以这里 move_penalty 必定为 1.0，损失为 0
        gate_penalty = F_cz_val ** num_cz
        move_penalty = 1.0  # 因为原始设定的 F_trans = 1.0
        deco_penalty = math.exp(-idle / T_eff)
        
        final_fid = gate_penalty * move_penalty * deco_penalty
        
        # 计算影响占比 (Logarithmic Loss Contribution)
        loss_gate = -math.log(gate_penalty) if gate_penalty > 0 else 0
        loss_move = -math.log(move_penalty) if move_penalty > 0 else 0
        loss_deco = -math.log(deco_penalty) if deco_penalty > 0 else 0
        total_loss = loss_gate + loss_move + loss_deco
        
        if total_loss > 0:
            pct_gate = loss_gate / total_loss
            pct_move = loss_move / total_loss
            pct_deco = loss_deco / total_loss
        else:
            pct_gate = pct_move = pct_deco = 0
            
        data_rows.append([
            circ, num_q, engine, 
            final_fid,
            gate_penalty, pct_gate,
            deco_penalty, pct_deco,
            move_penalty, pct_move,
            idle, num_cz
        ])

# ================= 写入 Excel =================
wb = Workbook()
ws = wb.active
ws.title = '保真度归因分析'

hdr_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
center = Alignment(horizontal='center', vertical='center')
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

headers = [
    '电路名称', '量子比特', '编译器', '最终保真度',
    '门错误惩罚 (f_cz^m)', '门错误占比',
    '退相干惩罚 (exp)', '退相干占比',
    '光镊抓放惩罚', '光镊错误占比',
    'T_idle (微秒)', '双量子门数'
]

for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = thin
    
data_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', size=10, bold=True)
gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

r_idx = 2
for row_data in data_rows:
    circ, q, eng, fid, pG, pctG, pD, pctD, pM, pctM, idle, n_cz = row_data
    
    ws.cell(row=r_idx, column=1, value=circ).font = bold_font
    ws.cell(row=r_idx, column=2, value=q).font = data_font
    ws.cell(row=r_idx, column=3, value=eng).font = bold_font
    if 'Baseline' in eng:
        for c in range(1, 13): ws.cell(row=r_idx, column=c).fill = gray_fill
    
    # 保真度
    ws.cell(row=r_idx, column=4, value=fid).number_format = '0.0000'
    ws.cell(row=r_idx, column=4).font = bold_font
    
    # 门惩罚 & 占比
    ws.cell(row=r_idx, column=5, value=pG).number_format = '0.0000'
    ws.cell(row=r_idx, column=6, value=pctG).number_format = '0.0%'
    ws.cell(row=r_idx, column=5).font = data_font
    ws.cell(row=r_idx, column=6).font = Font(name='微软雅黑', color='C00000', bold=True) # 红色高亮占比
    
    # 退相干惩罚 & 占比
    ws.cell(row=r_idx, column=7, value=pD).number_format = '0.0000'
    ws.cell(row=r_idx, column=8, value=pctD).number_format = '0.0%'
    ws.cell(row=r_idx, column=7).font = data_font
    ws.cell(row=r_idx, column=8).font = Font(name='微软雅黑', color='0070C0', bold=True) # 蓝色高亮占比
    
    # 移动惩罚 & 占比
    ws.cell(row=r_idx, column=9, value=pM).number_format = '0.0000'
    ws.cell(row=r_idx, column=10, value=pctM).number_format = '0.0%'
    ws.cell(row=r_idx, column=9).font = data_font
    ws.cell(row=r_idx, column=10).font = data_font
    
    # 基础数值
    ws.cell(row=r_idx, column=11, value=round(idle, 1)).font = data_font
    ws.cell(row=r_idx, column=12, value=n_cz).font = data_font
    
    for c in range(1, 13):
        ws.cell(row=r_idx, column=c).alignment = center
        ws.cell(row=r_idx, column=c).border = thin
        
    r_idx += 1

# 列宽设置
widths = [14, 10, 18, 12, 18, 12, 16, 12, 14, 12, 14, 12]
for col, w in zip('ABCDEFGHIJKL', widths):
    ws.column_dimensions[col].width = w

out_path = r'e:\coding\DasAtom_reading\DasAtom\fidelity_analysis.xlsx'
wb.save(out_path)
print(f'Done! Saved to {out_path}')
