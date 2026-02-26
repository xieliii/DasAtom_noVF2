"""Generate comparison report and save to markdown file."""
from openpyxl import load_workbook

base_file = r'e:\coding\DasAtom_reading\DasAtom_Origin\res\baseline_benchmark\Rb2Re4\h2h_bench_summary.xlsx'
dual_file = r'e:\coding\DasAtom_reading\DasAtom\res\dual_benchmark\Rb2Re4\h2h_bench_summary.xlsx'
output_file = r'C:\Users\14897\.gemini\antigravity\brain\61832c88-3d8e-4a79-9e30-3de50ebd81b4\benchmark_report.md'


def read_summary(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    data = []
    for row in rows[1:]:
        d = dict(zip(headers, row))
        if d.get('QASM File') and str(d['QASM File']).endswith('.qasm'):
            data.append(d)
    wb.close()
    return data


base_data = read_summary(base_file)
dual_data = read_summary(dual_file)
dual_map = {d['QASM File']: d for d in dual_data}

lines = []
lines.append('# Benchmark Report: Original VF2 vs MCTS + Force-Directed')
lines.append('')
lines.append('| Circuit | Qubits | Dist_Base | Dist_Dual | Dist_Imp | Fid_Base | Fid_Dual | Time_B(s) | Time_D(s) |')
lines.append('|---------|--------|-----------|-----------|----------|----------|----------|-----------|-----------|')

total_dist_imp = []
total_fb = []
total_fd = []
total_tb = []
total_td = []
dist_wins = 0
fid_wins = 0

for b in base_data:
    fname = b['QASM File']
    if fname not in dual_map:
        continue
    d = dual_map[fname]

    c = fname.replace('.qasm', '')
    q = int(b['Num Qubits'])
    db = float(b['Total Move Distance'])
    dd = float(d['Total Move Distance'])
    fb = float(b['Fidelity'])
    fd = float(d['Fidelity'])
    tb = float(b['Elapsed Time (s)'])
    td = float(d['Elapsed Time (s)'])

    if db > 0:
        di = (db - dd) / db * 100
        imp_s = '{:+.1f}%'.format(di)
        total_dist_imp.append(di)
    else:
        imp_s = 'N/A'

    total_fb.append(fb)
    total_fd.append(fd)
    total_tb.append(tb)
    total_td.append(td)

    if dd < db:
        dist_wins += 1
    if fd > fb:
        fid_wins += 1

    lines.append('| {} | {} | {:.1f} | {:.1f} | {} | {:.6f} | {:.6f} | {:.2f} | {:.2f} |'.format(
        c, q, db, dd, imp_s, fb, fd, tb, td))

n = len(total_fb)
lines.append('')
lines.append('## Summary')
lines.append('')
if total_dist_imp:
    lines.append('- **Avg distance improvement**: {:+.1f}%'.format(sum(total_dist_imp) / len(total_dist_imp)))
lines.append('- **Avg fidelity**: baseline={:.6f}, dual={:.6f}'.format(sum(total_fb)/n, sum(total_fd)/n))
lines.append('- **Avg compile time**: baseline={:.3f}s, dual={:.3f}s'.format(sum(total_tb)/n, sum(total_td)/n))
lines.append('- **Circuits tested**: {}'.format(n))
lines.append('- **Dual wins on distance**: {}/{}'.format(dist_wins, n))
lines.append('- **Dual wins on fidelity**: {}/{}'.format(fid_wins, n))

report = '\n'.join(lines)
print(report)

with open(output_file, 'w', encoding='utf-8') as f:
    f.write(report)

print()
print('Report saved to:', output_file)
