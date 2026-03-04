"""
generate_ratio_report.py
========================
学术风格的倍数报告生成器。
将 Baseline (VF2) 与 Dual (MCTS + Force-Directed) 的基准测试结果，
转化为更直观的 加速比 / 缩减率 / 保真度倍数 表格，用于论文和组会展示。

输出:
  1. 终端打印 Markdown 表格
  2. 保存 academic_ratio_report.xlsx
"""

import sys
import os

# Windows 终端 UTF-8 兼容
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ──────────────────────────── 文件路径 ────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(BASE_DIR)

DUAL_SUMMARY = os.path.join(BASE_DIR, "res", "dual_benchmark", "Rb2Re4", "h2h_bench_summary.xlsx")
BASE_SUMMARY = os.path.join(ROOT_DIR, "DasAtom_Origin", "res", "baseline_benchmark", "Rb2Re4", "h2h_bench_summary.xlsx")
FIDELITY_ANALYSIS = os.path.join(BASE_DIR, "fidelity_analysis.xlsx")
OUTPUT_XLSX = os.path.join(BASE_DIR, "academic_ratio_report.xlsx")


# ──────────────────────────── 数据读取 ────────────────────────────

def read_h2h_summary(path: str) -> dict:
    """读取 h2h_bench_summary.xlsx，返回 {circuit_name: row_dict}"""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    data = {}
    for row in rows[1:]:
        d = dict(zip(headers, row))
        fname = d.get("QASM File")
        if fname and str(fname).endswith(".qasm"):
            data[fname.replace(".qasm", "")] = d
    wb.close()
    return data


def read_fidelity_analysis(path: str) -> dict:
    """
    读取 fidelity_analysis.xlsx，提取退相干保真度 F_deco。
    返回 {circuit_name: {"baseline": F_deco_base, "dual": F_deco_dual}}
    """
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    result = {}
    for row in rows[1:]:
        d = dict(zip(headers, row))
        circuit = d.get("电路名称") or d.get(headers[0])
        compiler = d.get("编译器") or d.get(headers[2])
        # 退相干惩罚 (exp) 列 —— 这就是 F_deco
        f_deco = d.get("退相干惩罚 (exp)") or d.get(headers[6])
        if circuit is None or compiler is None or f_deco is None:
            continue
        try:
            f_deco = float(f_deco)
        except (ValueError, TypeError):
            continue
        if circuit not in result:
            result[circuit] = {}
        if "Baseline" in str(compiler) or "原版" in str(compiler) or "VF2" in str(compiler):
            result[circuit]["baseline"] = f_deco
        elif "Dual" in str(compiler) or "纯血" in str(compiler) or "新版" in str(compiler):
            result[circuit]["dual"] = f_deco
    wb.close()
    return result


# ──────────────────────────── 指标计算 ────────────────────────────

def calc_time_speedup(t_base: float, t_dual: float) -> tuple:
    """编译时间加速比。返回 (数值, 格式化字符串)"""
    if t_dual <= 0:
        return (float("inf"), "∞")
    ratio = t_base / t_dual
    return (ratio, f"{ratio:.1f}x")


def calc_dist_reduction(d_base: float, d_dual: float) -> tuple:
    """距离缩减率。返回 (数值, 格式化字符串)"""
    if d_base == 0 and d_dual == 0:
        return (0.0, "Optimal (0.0)")
    if d_base == 0:
        return (float("-inf"), "N/A")
    rate = (d_base - d_dual) / d_base * 100
    return (rate, f"{rate:+.1f}%")


def calc_fdeco_ratio(f_base: float, f_dual: float) -> tuple:
    """退相干保真度倍数。返回 (数值, 格式化字符串)"""
    if f_base <= 0:
        return (float("inf"), "∞")
    ratio = f_dual / f_base
    return (ratio, f"{ratio:.2f}x")


def judge_win(dist_reduction_val: float, fdeco_ratio_val: float, dist_str: str) -> str:
    """综合胜负判定"""
    if dist_str == "Optimal (0.0)":
        return "⚪ Tie"
    if dist_reduction_val > 0 and fdeco_ratio_val > 1.0:
        return "🟢 Dual Win"
    elif dist_reduction_val < 0 and fdeco_ratio_val < 1.0:
        return "🔴 Base Win"
    else:
        return "🟡 Mixed"


# ──────────────────────────── 主逻辑 ────────────────────────────

def main():
    print("=" * 72)
    print("  📊 Academic Ratio Report Generator")
    print("  Dual-Engine (MCTS + FD) vs Baseline (VF2)")
    print("=" * 72)
    print()

    # 读取数据
    dual_data = read_h2h_summary(DUAL_SUMMARY)
    base_data = read_h2h_summary(BASE_SUMMARY)
    fdeco_data = read_fidelity_analysis(FIDELITY_ANALYSIS)

    # 按电路名排序（先按 qubits 再按名字）
    all_circuits = sorted(
        set(dual_data.keys()) & set(base_data.keys()),
        key=lambda c: (int(base_data[c].get("Num Qubits", 0)), c)
    )

    print(f"  ✅ 找到 {len(all_circuits)} 个匹配电路\n")

    # ───── 计算所有指标 ─────
    results = []
    for circuit in all_circuits:
        bd = base_data[circuit]
        dd = dual_data[circuit]

        qubits = int(bd.get("Num Qubits", 0))
        t_base = float(bd.get("Elapsed Time (s)", 0))
        t_dual = float(dd.get("Elapsed Time (s)", 0))
        d_base = float(bd.get("Total Move Distance", 0))
        d_dual = float(dd.get("Total Move Distance", 0))

        speedup_val, speedup_str = calc_time_speedup(t_base, t_dual)
        dist_val, dist_str = calc_dist_reduction(d_base, d_dual)

        # 退相干保真度
        fdeco = fdeco_data.get(circuit, {})
        f_base = fdeco.get("baseline", None)
        f_dual = fdeco.get("dual", None)
        if f_base is not None and f_dual is not None:
            fdeco_val, fdeco_str = calc_fdeco_ratio(f_base, f_dual)
        else:
            fdeco_val, fdeco_str = (1.0, "N/A")

        win = judge_win(dist_val, fdeco_val, dist_str)

        # 强调标记
        speedup_display = speedup_str
        if qubits >= 16 and speedup_val >= 2.0:
            speedup_display = f"🚀 {speedup_str}"

        dist_display = dist_str
        if dist_val > 50:
            dist_display = f"🎯 {dist_str}"

        results.append({
            "circuit": circuit,
            "qubits": qubits,
            "speedup_val": speedup_val,
            "speedup_str": speedup_display,
            "dist_val": dist_val,
            "dist_str": dist_display,
            "dist_str_raw": dist_str,
            "fdeco_val": fdeco_val,
            "fdeco_str": fdeco_str,
            "win": win,
        })

    # ───── Markdown 表格输出 ─────
    md_header = "| Circuit | Qubits | Time_Speedup | Dist_Reduction | F_deco_Ratio | Win/Loss |"
    md_sep =    "|---------|--------|:------------:|:--------------:|:------------:|:--------:|"
    print(md_header)
    print(md_sep)
    for r in results:
        print(f"| {r['circuit']} | {r['qubits']} | {r['speedup_str']} | {r['dist_str']} | {r['fdeco_str']} | {r['win']} |")

    # ───── 高光总结 ─────
    print()
    print("=" * 72)
    print("  🏆 Highlight Summary")
    print("=" * 72)

    # Top 3 加速比
    top_speed = sorted(results, key=lambda r: r["speedup_val"], reverse=True)[:3]
    print("\n  ⚡ Top 3 编译加速比 (Compilation Speedup):")
    for i, r in enumerate(top_speed, 1):
        print(f"     {i}. {r['circuit']} ({r['qubits']}Q) — {r['speedup_str']}")

    # Top 3 距离缩减
    top_dist = sorted(results, key=lambda r: r["dist_val"], reverse=True)[:3]
    print("\n  📏 Top 3 距离缩减率 (Distance Reduction):")
    for i, r in enumerate(top_dist, 1):
        print(f"     {i}. {r['circuit']} ({r['qubits']}Q) — {r['dist_str_raw']}")

    # 大电路平均加速比 (>= 16 qubits)
    large = [r for r in results if r["qubits"] >= 16]
    if large:
        avg_speedup = sum(r["speedup_val"] for r in large) / len(large)
        print(f"\n  🔬 大电路 (Qubits ≥ 16) 平均加速比: {avg_speedup:.1f}x  ({len(large)} circuits)")

    # 总体胜率
    wins = sum(1 for r in results if "Dual Win" in r["win"])
    ties = sum(1 for r in results if "Tie" in r["win"])
    total = len(results)
    print(f"\n  📈 总体战绩: 🟢 Win {wins} / ⚪ Tie {ties} / Total {total}")
    print("=" * 72)

    # ───── 写入 Excel ─────
    write_excel(results)
    print(f"\n  💾 Excel 报告已保存: {OUTPUT_XLSX}")
    print()


# ──────────────────────────── Excel 输出 ────────────────────────────

def write_excel(results: list):
    """生成格式精美的 academic_ratio_report.xlsx"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ratio Report"

    # 样式定义
    header_font = Font(name="Consolas", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Consolas", size=10)
    data_align = Alignment(horizontal="center", vertical="center")
    win_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    tie_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    highlight_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # 表头
    headers = ["Circuit", "Qubits", "Time_Speedup", "Dist_Reduction", "F_deco_Ratio", "Win/Loss"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 数据行
    for row_idx, r in enumerate(results, 2):
        values = [
            r["circuit"],
            r["qubits"],
            r["speedup_str"],
            r["dist_str"],
            r["fdeco_str"],
            r["win"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

        # 行级颜色
        win_str = r["win"]
        if "Dual Win" in win_str:
            fill = win_fill
        elif "Base Win" in win_str:
            fill = loss_fill
        elif "Tie" in win_str:
            fill = tie_fill
        else:
            fill = None

        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

        # 高亮加速比 > 2.0x 的大电路
        if r["qubits"] >= 16 and r["speedup_val"] >= 2.0:
            ws.cell(row=row_idx, column=3).fill = highlight_fill
            ws.cell(row=row_idx, column=3).font = Font(name="Consolas", size=10, bold=True, color="9C5700")

        # 高亮距离缩减 > 50%
        if r["dist_val"] > 50:
            ws.cell(row=row_idx, column=4).fill = highlight_fill
            ws.cell(row=row_idx, column=4).font = Font(name="Consolas", size=10, bold=True, color="9C5700")

    # ───── 高光总结 Sheet ─────
    ws_summary = wb.create_sheet("Highlight Summary")
    summary_header_font = Font(name="Consolas", bold=True, size=12, color="2F5496")
    summary_font = Font(name="Consolas", size=10)

    row = 1
    ws_summary.cell(row=row, column=1, value="🏆 Highlight Summary").font = Font(name="Consolas", bold=True, size=14, color="2F5496")
    row += 2

    # Top 3 加速比
    ws_summary.cell(row=row, column=1, value="⚡ Top 3 Compilation Speedup").font = summary_header_font
    row += 1
    top_speed = sorted(results, key=lambda r: r["speedup_val"], reverse=True)[:3]
    for i, r in enumerate(top_speed, 1):
        ws_summary.cell(row=row, column=1, value=f"{i}. {r['circuit']} ({r['qubits']}Q)").font = summary_font
        ws_summary.cell(row=row, column=2, value=r["speedup_str"]).font = Font(name="Consolas", size=10, bold=True)
        row += 1

    row += 1
    # Top 3 距离缩减
    ws_summary.cell(row=row, column=1, value="📏 Top 3 Distance Reduction").font = summary_header_font
    row += 1
    top_dist = sorted(results, key=lambda r: r["dist_val"], reverse=True)[:3]
    for i, r in enumerate(top_dist, 1):
        ws_summary.cell(row=row, column=1, value=f"{i}. {r['circuit']} ({r['qubits']}Q)").font = summary_font
        ws_summary.cell(row=row, column=2, value=r["dist_str_raw"]).font = Font(name="Consolas", size=10, bold=True)
        row += 1

    row += 1
    # 大电路平均加速比
    large = [r for r in results if r["qubits"] >= 16]
    if large:
        avg_sp = sum(r["speedup_val"] for r in large) / len(large)
        ws_summary.cell(row=row, column=1, value="🔬 大电路 (Qubits ≥ 16) 平均加速比").font = summary_header_font
        ws_summary.cell(row=row, column=2, value=f"{avg_sp:.1f}x").font = Font(name="Consolas", size=12, bold=True, color="C00000")
        row += 1

    row += 1
    wins = sum(1 for r in results if "Dual Win" in r["win"])
    ties = sum(1 for r in results if "Tie" in r["win"])
    ws_summary.cell(row=row, column=1, value="📈 总体战绩").font = summary_header_font
    ws_summary.cell(row=row, column=2, value=f"Win {wins} / Tie {ties} / Total {len(results)}").font = Font(name="Consolas", size=10, bold=True)

    # 列宽调整
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws_summary.column_dimensions["A"].width = 40
    ws_summary.column_dimensions["B"].width = 20

    wb.save(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
