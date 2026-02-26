"""
generate_fidelity_excel.py — 生成保真度对比 Excel 报告

内容：
1. Sheet1: 总体对比表 (Baseline vs Dual, 百分比提升)
2. Sheet2: 保真度分解 (F_deco / F_cz / F_trans 各组分)
3. Sheet3: qft_8 回归深度分析
"""

import os, math, json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

# ============ 物理参数 ============
T_cz = 0.2
T_eff = 1.5e6
T_trans = 20
Move_speed = 0.55
F_cz_val = 0.995
AOD_width = 3
AOD_height = 3

# ============ 路径 ============
base_dir = r'e:\coding\DasAtom_reading\DasAtom_Origin\res\baseline_benchmark\Rb2Re4'
dual_dir = r'e:\coding\DasAtom_reading\DasAtom\res\dual_benchmark\Rb2Re4'

circuits = [
    'linear_6', 'qft_6', 'qv_6', 'random_6', 'star_6',
    'linear_8', 'qft_8', 'qv_8', 'random_8', 'star_8',
    'linear_12', 'qft_12', 'qv_12', 'random_12', 'star_12',
    'linear_16', 'qft_16', 'qv_16', 'random_16', 'star_16',
    'linear_20', 'qft_20', 'qv_20', 'random_20', 'star_20'
]

def extract_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    d = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] and isinstance(row[0], str):
            d[row[0].strip()] = row[1]
    wb.close()
    return d

def safe_get(d, *keys):
    for k in keys:
        if k in d:
            return d[k]
    return None

# ============ 收集数据 ============
all_data = []
for circ in circuits:
    fb = os.path.join(base_dir, f'{circ}.qasm_rb2.xlsx')
    fd = os.path.join(dual_dir, f'{circ}.qasm_rb2.xlsx')
    
    base_d = extract_xlsx(fb) if os.path.exists(fb) else {}
    dual_d = extract_xlsx(fd) if os.path.exists(fd) else {}
    
    if not base_d or not dual_d:
        continue
    
    num_q = safe_get(base_d, 'Number of qubits (num_q)', 'Num qubits') or 0
    gate_num_b = safe_get(base_d, 'Number of CZ gates') or 0
    gate_num_d = safe_get(dual_d, 'Number of CZ gates') or 0
    
    dist_b = safe_get(base_d, 'Total Move Distance', 'Total move distance') or 0
    dist_d = safe_get(dual_d, 'Total Move Distance', 'Total move distance') or 0
    
    fid_b = safe_get(base_d, 'Fidelity') or 0
    fid_d = safe_get(dual_d, 'Fidelity') or 0
    
    idle_b = safe_get(base_d, 'Idle time') or 0
    idle_d = safe_get(dual_d, 'Idle time') or 0
    
    t_total_b = safe_get(base_d, 'Total_T (from fidelity calc)') or 0
    t_total_d = safe_get(dual_d, 'Total_T (from fidelity calc)') or 0
    
    # 分解保真度
    f_cz_b = F_cz_val ** gate_num_b if gate_num_b else 1.0
    f_cz_d = F_cz_val ** gate_num_d if gate_num_d else 1.0
    f_deco_b = math.exp(-idle_b / T_eff) if idle_b else 1.0
    f_deco_d = math.exp(-idle_d / T_eff) if idle_d else 1.0
    
    all_data.append({
        'circ': circ, 'num_q': num_q,
        'gate_num_b': gate_num_b, 'gate_num_d': gate_num_d,
        'dist_b': dist_b, 'dist_d': dist_d,
        'fid_b': fid_b, 'fid_d': fid_d,
        'idle_b': idle_b, 'idle_d': idle_d,
        't_total_b': t_total_b, 't_total_d': t_total_d,
        'f_cz_b': f_cz_b, 'f_cz_d': f_cz_d,
        'f_deco_b': f_deco_b, 'f_deco_d': f_deco_d,
    })

# ============ 样式 ============
wb = Workbook()

hdr_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
hdr_fill2 = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
data_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', size=10, bold=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

green_font = Font(name='微软雅黑', size=10, bold=True, color='006100')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_font = Font(name='微软雅黑', size=10, bold=True, color='9C0006')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

def style_header(ws, row, headers, fill=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hdr_font
        cell.fill = fill or hdr_fill
        cell.alignment = center
        cell.border = thin

def style_data_cell(ws, row, col, value, fmt=None, font=None, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or data_font
    cell.alignment = center
    cell.border = thin
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    return cell

# ===============================================================
# Sheet 1: 总体对比表
# ===============================================================
ws1 = wb.active
ws1.title = '保真度对比总表'

headers1 = [
    '电路', 'Qubits', 'CZ门数',
    'Baseline\n保真度', 'Dual\n保真度', '保真度\n提升(%)',
    'Baseline\n距离', 'Dual\n距离', '距离\n缩减(%)',
    '主导\n保真度', 'Dual是否\n胜出'
]
style_header(ws1, 1, headers1)

r = 2
for dd in all_data:
    fid_pct = ((dd['fid_d'] - dd['fid_b']) / dd['fid_b'] * 100) if dd['fid_b'] > 0 else 0
    dist_pct = ((dd['dist_b'] - dd['dist_d']) / dd['dist_b'] * 100) if dd['dist_b'] > 0 else 0
    
    has_diff = abs(dd['fid_d'] - dd['fid_b']) > 1e-9
    dual_wins = dd['fid_d'] > dd['fid_b'] + 1e-9
    
    style_data_cell(ws1, r, 1, dd['circ'], font=bold_font)
    style_data_cell(ws1, r, 2, dd['num_q'])
    style_data_cell(ws1, r, 3, dd['gate_num_b'])
    
    # 保真度列 — 核心对比
    style_data_cell(ws1, r, 4, dd['fid_b'], fmt='0.000000')
    style_data_cell(ws1, r, 5, dd['fid_d'], fmt='0.000000')
    
    # 保真度提升百分比
    if has_diff:
        pct_font = green_font if fid_pct > 0 else red_font
        pct_fill = green_fill if fid_pct > 0 else red_fill
        style_data_cell(ws1, r, 6, fid_pct / 100, fmt='0.00%', font=pct_font, fill=pct_fill)
    else:
        style_data_cell(ws1, r, 6, 0, fmt='0.00%', fill=yellow_fill)
    
    # 距离
    style_data_cell(ws1, r, 7, dd['dist_b'], fmt='0.00')
    style_data_cell(ws1, r, 8, dd['dist_d'], fmt='0.00')
    
    if dd['dist_b'] > 0:
        dist_font = green_font if dist_pct > 0 else red_font
        dist_fill = green_fill if dist_pct > 0 else red_fill
        style_data_cell(ws1, r, 9, dist_pct / 100, fmt='0.0%', font=dist_font, fill=dist_fill)
    else:
        style_data_cell(ws1, r, 9, 0, fmt='0.0%', fill=yellow_fill)
    
    # 主导保真度
    style_data_cell(ws1, r, 10, '退相干 F_deco' if has_diff else '无差异')
    
    # 胜负
    if not has_diff:
        style_data_cell(ws1, r, 11, '平局', fill=yellow_fill)
    elif dual_wins:
        style_data_cell(ws1, r, 11, '✓ 胜', font=green_font, fill=green_fill)
    else:
        style_data_cell(ws1, r, 11, '✗ 败', font=red_font, fill=red_fill)
    
    r += 1

# 汇总行
r += 1
style_data_cell(ws1, r, 1, '汇总', font=Font(name='微软雅黑', bold=True, size=12))
fid_vals = [(d['fid_b'], d['fid_d']) for d in all_data]
avg_b = sum(f[0] for f in fid_vals) / len(fid_vals)
avg_d = sum(f[1] for f in fid_vals) / len(fid_vals)
avg_pct = (avg_d - avg_b) / avg_b * 100

style_data_cell(ws1, r, 4, avg_b, fmt='0.000000', font=bold_font)
style_data_cell(ws1, r, 5, avg_d, fmt='0.000000', font=bold_font)
style_data_cell(ws1, r, 6, avg_pct / 100, fmt='0.00%', font=green_font, fill=green_fill)

# 有差异电路的平均
interesting = [d for d in all_data if abs(d['fid_d'] - d['fid_b']) > 1e-9]
if interesting:
    wins = sum(1 for d in interesting if d['fid_d'] > d['fid_b'])
    style_data_cell(ws1, r, 10, f'胜: {wins}/{len(interesting)}', font=bold_font)

# 列宽
widths1 = [14, 8, 8, 14, 14, 12, 12, 12, 12, 14, 10]
for col_letter, w in zip('ABCDEFGHIJK', widths1):
    ws1.column_dimensions[col_letter].width = w

# ===============================================================
# Sheet 2: 保真度组分分解
# ===============================================================
ws2 = wb.create_sheet('保真度分解')

headers2 = [
    '电路', 'Qubits',
    'Base\nF_cz', 'Dual\nF_cz', 'ΔF_cz',
    'Base\nF_deco', 'Dual\nF_deco', 'F_deco\n提升(%)',
    'Base\nt_idle(μs)', 'Dual\nt_idle(μs)', 't_idle\n减少(%)',
    'Base\n距离', 'Dual\n距离', '距离\n缩减(%)',
]
style_header(ws2, 1, headers2, fill=hdr_fill2)

r = 2
for dd in all_data:
    has_diff = abs(dd['fid_d'] - dd['fid_b']) > 1e-9
    if not has_diff:
        continue
    
    style_data_cell(ws2, r, 1, dd['circ'], font=bold_font)
    style_data_cell(ws2, r, 2, dd['num_q'])
    
    # F_cz
    style_data_cell(ws2, r, 3, dd['f_cz_b'], fmt='0.000000')
    style_data_cell(ws2, r, 4, dd['f_cz_d'], fmt='0.000000')
    style_data_cell(ws2, r, 5, dd['f_cz_d'] - dd['f_cz_b'], fmt='0.000000')
    
    # F_deco
    style_data_cell(ws2, r, 6, dd['f_deco_b'], fmt='0.000000')
    style_data_cell(ws2, r, 7, dd['f_deco_d'], fmt='0.000000')
    
    deco_pct = (dd['f_deco_d'] - dd['f_deco_b']) / dd['f_deco_b'] * 100 if dd['f_deco_b'] > 0 else 0
    pct_font = green_font if deco_pct > 0 else red_font
    pct_fill = green_fill if deco_pct > 0 else red_fill
    style_data_cell(ws2, r, 8, deco_pct / 100, fmt='0.000%', font=pct_font, fill=pct_fill)
    
    # t_idle
    style_data_cell(ws2, r, 9, round(dd['idle_b'], 1), fmt='#,##0.0')
    style_data_cell(ws2, r, 10, round(dd['idle_d'], 1), fmt='#,##0.0')
    idle_pct = (dd['idle_b'] - dd['idle_d']) / dd['idle_b'] * 100 if dd['idle_b'] > 0 else 0
    idle_font = green_font if idle_pct > 0 else red_font
    idle_fill = green_fill if idle_pct > 0 else red_fill
    style_data_cell(ws2, r, 11, idle_pct / 100, fmt='0.0%', font=idle_font, fill=idle_fill)
    
    # 距离
    style_data_cell(ws2, r, 12, dd['dist_b'], fmt='0.00')
    style_data_cell(ws2, r, 13, dd['dist_d'], fmt='0.00')
    dist_pct = (dd['dist_b'] - dd['dist_d']) / dd['dist_b'] * 100 if dd['dist_b'] > 0 else 0
    dist_font = green_font if dist_pct > 0 else red_font
    dist_fill = green_fill if dist_pct > 0 else red_fill
    style_data_cell(ws2, r, 14, dist_pct / 100, fmt='0.0%', font=dist_font, fill=dist_fill)
    
    r += 1

widths2 = [14, 8, 10, 10, 10, 12, 12, 12, 14, 14, 12, 12, 12, 12]
for col_letter, w in zip('ABCDEFGHIJKLMN', widths2):
    ws2.column_dimensions[col_letter].width = w

# ===============================================================
# Sheet 3: qft_8 回归深度分析
# ===============================================================
ws3 = wb.create_sheet('qft_8 回归分析')

# 标题
ws3.merge_cells('A1:H1')
title_cell = ws3.cell(row=1, column=1, value='qft_8 回归案例深度分析 — Dual保真度低于Baseline的唯一案例')
title_cell.font = Font(name='微软雅黑', bold=True, size=14, color='9C0006')
title_cell.alignment = center
title_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# 对比数据
data_items = [
    ('指标', 'Baseline (VF2)', 'Dual (MCTS+力导向)', '差异', '方向'),
    ('CZ门数', 68, 68, 0, '相同'),
    ('Total Move Distance', 24.90, 28.42, '+3.51', '❌ Dual更差'),
    ('t_idle (μs)', None, None, None, None),
    ('F_deco (退相干保真度)', None, None, None, None),
    ('F_cz (CZ门保真度)', None, None, None, '相同'),
    ('最终 Fidelity', 0.709432, 0.709105, '-0.000327', '❌ Dual更差'),
    ('保真度百分比变化', '', '', '-0.046%', '❌'),
    ('', '', '', '', ''),
    ('Per-qubit 距离之和', 12.129, 9.472, '-2.657', '✓ Dual更好'),
    ('需移动的 qubit 数', 8, 6, '-2', '✓ Dual更好'),
    ('', '', '', '', ''),
    ('矛盾关键:', '', '', '', ''),
    ('Dual per-qubit距离更短(9.47<12.13)', '', '', '', ''),
    ('但QuantumRouter处理后总距离反而更长!', '', '', '', ''),
]

# qft_8 数据补全
qft8_data = [d for d in all_data if d['circ'] == 'qft_8']
if qft8_data:
    q8 = qft8_data[0]
    data_items[3] = ('t_idle (μs)', round(q8['idle_b'], 1), round(q8['idle_d'], 1), 
                      f"+{round(q8['idle_d'] - q8['idle_b'], 1)}", '❌ Dual更差')
    data_items[4] = ('F_deco (退相干保真度)', round(q8['f_deco_b'], 6), round(q8['f_deco_d'], 6),
                      round(q8['f_deco_d'] - q8['f_deco_b'], 6), '❌ Dual更差')
    data_items[5] = ('F_cz (CZ门保真度)', round(q8['f_cz_b'], 6), round(q8['f_cz_d'], 6), 0, '相同')

r = 3
for item in data_items:
    for c, val in enumerate(item, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        if r == 3:
            cell.font = hdr_font
            cell.fill = hdr_fill
        else:
            cell.font = data_font
            if isinstance(val, str) and '❌' in val:
                cell.font = red_font
                cell.fill = red_fill
            elif isinstance(val, str) and '✓' in val:
                cell.font = green_font
                cell.fill = green_fill
        cell.alignment = center
        cell.border = thin
    r += 1

# 根因分析
r += 2
ws3.merge_cells(f'A{r}:H{r}')
ws3.cell(row=r, column=1, value='根因分析').font = Font(name='微软雅黑', bold=True, size=13)
r += 1

reasons = [
    '1. Dual的embedding: qubit的移动方向更"交叉"',
    '   例: q2 (0,0)→(1,2) 和 q3 (0,2)→(1,0)，方向互斥',
    '2. QuantumRouter的compatible_2D检测到更多冲突(violations)',
    '   → 需要分成更多步骤',
    '   → 每步取max_dis(最远qubit距离)作为时间开销',
    '   → 累积后总距离 28.42 > 24.90',
    '',
    '3. 根本原因: analytical_placer.py 的力导向模型',
    '   只考虑: ① 门引力(拉近有交互的qubit) ② 惯性力(锚定上一层位置)',
    '   不考虑: ③ AOD路由约束 (移动方向是否兼容)',
    '',
    '4. 改进方案:',
    '   a) 贪心吸附时考虑移动方向兼容性',
    '   b) MCTS目标函数纳入路由成本预估',
    '   c) 力导向增加前瞻(看下一层需求)',
    '   d) 小电路(≤10q)增加MCTS迭代或fallback VF2',
]

for reason in reasons:
    ws3.cell(row=r, column=1, value=reason).font = data_font
    ws3.merge_cells(f'A{r}:H{r}')
    r += 1

widths3 = [30, 18, 22, 14, 16, 10, 10, 10]
for col_letter, w in zip('ABCDEFGH', widths3):
    ws3.column_dimensions[col_letter].width = w

# ============ 保存 ============
out_path = r'e:\coding\DasAtom_reading\DasAtom\fidelity_comparison.xlsx'
wb.save(out_path)
print(f'✓ Excel报告已生成: {out_path}')
print(f'  Sheet1: 保真度对比总表 ({len(all_data)}个电路)')
print(f'  Sheet2: 保真度分解 ({len(interesting)}个有差异电路)')
print(f'  Sheet3: qft_8 回归分析')
