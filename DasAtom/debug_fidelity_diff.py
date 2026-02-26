"""
debug_fidelity_diff.py — 深入分析 Baseline vs Dual 保真度差异的来源

读取所有 per-circuit xlsx 结果文件，提取关键保真度组成成分进行对比。
保真度公式:  F = exp(-t_idle/T_eff) * F_cz^gate_num * F_trans^num_trans
  - CZ门保真度 (F_cz^gate_num): 由 CZ 门数量决定 (两边 gate_num 相同? 要验证)
  - 退相干保真度 (exp(-t_idle/T_eff)): 由 idle time 决定 (t_idle = num_q * t_total - gate_num * T_cz)
  - 传输保真度 (F_trans^num_trans): F_trans=1，所以始终为1

t_total = len(parallel_gates)*T_cz + sum_over_moves(4*T_trans + max_dis/Move_speed)
t_idle = num_q * t_total - gate_num * T_cz

所以关键变量:
  1. gate_num (CZ门数) — 理论上两边相同
  2. len(parallel_gates) — 并行分组数(影响 t_total)
  3. all_movements 的数量和距离 — 移动步数和距离
  4. num_q — 量子比特数
"""

import os
import math
from openpyxl import load_workbook

T_cz = 0.2
T_eff = 1.5e6
T_trans = 20
Move_speed = 0.55
F_cz_val = 0.995

base_dir = r'e:\coding\DasAtom_reading\DasAtom_Origin\res\baseline_benchmark\Rb2Re4'
dual_dir = r'e:\coding\DasAtom_reading\DasAtom\res\dual_benchmark\Rb2Re4'

circuits = [
    'linear_6', 'qft_6', 'qv_6', 'random_6', 'star_6',
    'linear_8', 'qft_8', 'qv_8', 'random_8', 'star_8',
    'linear_12', 'qft_12', 'qv_12', 'random_12', 'star_12',
    'linear_16', 'qft_16', 'qv_16', 'random_16', 'star_16',
    'linear_20', 'qft_20', 'qv_20', 'random_20', 'star_20'
]

def extract_all_fields(path):
    """提取xlsx中所有键值对"""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    d = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] and isinstance(row[0], str):
            d[row[0].strip()] = row[1]
    wb.close()
    return d

def safe_get(d, *keys):
    for k in keys:
        if k in d:
            return d[k]
    return None

print("=" * 120)
print(f"{'电路':>12} | {'Engine':>10} | {'gate_num':>8} | {'PG_steps':>8} | {'MoveSteps':>9} | {'TotalDist':>9} | {'t_total':>10} | {'t_idle':>10} | {'F_deco':>10} | {'F_cz':>10} | {'F_final':>10}")
print("-" * 120)

diff_data = []

for circ in circuits:
    fb = os.path.join(base_dir, f'{circ}.qasm_rb2.xlsx')
    fd = os.path.join(dual_dir, f'{circ}.qasm_rb2.xlsx')
    
    row_data = {}
    
    for tag, fp in [('Base', fb), ('Dual', fd)]:
        if not os.path.exists(fp):
            continue
        d = extract_all_fields(fp)
        
        # 打印所有可用的键名，用于调试
        if circ == circuits[0] and tag == 'Base':
            print(f"\n[DEBUG] Available keys in {circ} ({tag}):")
            for k, v in d.items():
                print(f"  '{k}' = {v}")
            print()
        
        num_q = safe_get(d, 'Number of qubits (num_q)', 'Num qubits', 'num_q')
        gate_num = safe_get(d, 'Number of CZ gates', 'gate_num', 'CZ gates')
        
        # Parallel gate steps
        pg_steps = safe_get(d, 'Parallel gate steps', 'Number of parallel gate steps')
        
        # Movement steps
        move_steps = safe_get(d, 'Number of movement steps', 'Movement steps', 'num_move')
        
        # Total move distance
        total_dist = safe_get(d, 'Total Move Distance', 'Total move distance', 'all_move_dis')
        
        # t_total
        t_total = safe_get(d, 'Total_T (from fidelity calc)', 'Total T', 't_total')
        
        # t_idle
        t_idle = safe_get(d, 'Idle time', 't_idle')
        
        # Fidelity
        fidelity = safe_get(d, 'Fidelity')
        
        # num_trans
        num_trans = safe_get(d, 'Number of AOD transfers', 'num_trans')
        
        # 如果能算就算
        if gate_num and gate_num > 0:
            f_cz = F_cz_val ** gate_num
        else:
            f_cz = 1.0
            
        if t_idle and t_idle > 0:
            f_deco = math.exp(-t_idle / T_eff)
        else:
            f_deco = 1.0
        
        row_data[tag] = {
            'num_q': num_q, 'gate_num': gate_num, 'pg_steps': pg_steps,
            'move_steps': move_steps, 'total_dist': total_dist,
            't_total': t_total, 't_idle': t_idle,
            'f_deco': f_deco, 'f_cz': f_cz, 'fidelity': fidelity,
            'num_trans': num_trans,
        }
        
        print(f"{circ:>12} | {tag:>10} | {gate_num or 0:>8} | {pg_steps or 0:>8} | {move_steps or 0:>9} | {total_dist or 0:>9.2f} | {t_total or 0:>10.2f} | {t_idle or 0:>10.2f} | {f_deco:>10.6f} | {f_cz:>10.6f} | {fidelity or 0:>10.6f}")
    
    # 计算差异
    if 'Base' in row_data and 'Dual' in row_data:
        b = row_data['Base']
        d_data = row_data['Dual']
        
        fid_diff = (d_data['fidelity'] or 0) - (b['fidelity'] or 0)
        deco_diff = (d_data['f_deco'] or 1) - (b['f_deco'] or 1)
        cz_diff = (d_data['f_cz'] or 1) - (b['f_cz'] or 1)
        idle_diff = (d_data['t_idle'] or 0) - (b['t_idle'] or 0)
        dist_diff = (d_data['total_dist'] or 0) - (b['total_dist'] or 0)
        pg_diff = (d_data['pg_steps'] or 0) - (b['pg_steps'] or 0)
        
        diff_data.append({
            'circ': circ,
            'fid_diff': fid_diff, 'deco_diff': deco_diff, 'cz_diff': cz_diff,
            'idle_diff': idle_diff, 'dist_diff': dist_diff, 'pg_diff': pg_diff,
            'base_pg': b['pg_steps'], 'dual_pg': d_data['pg_steps'],
            'base_idle': b['t_idle'], 'dual_idle': d_data['t_idle'],
            'base_dist': b['total_dist'], 'dual_dist': d_data['total_dist'],
            'base_gate_num': b['gate_num'], 'dual_gate_num': d_data['gate_num'],
            'base_move_steps': b['move_steps'], 'dual_move_steps': d_data['move_steps'],
        })
    
    print("-" * 120)

# ============ 差异分析汇总 ============
print("\n\n" + "=" * 100)
print("差异分析汇总 (Dual - Base, 正值=Dual更大/更差)")
print("=" * 100)
print(f"{'电路':>12} | {'Fid_Diff':>10} | {'Deco_Diff':>10} | {'CZ_Diff':>10} | {'Idle_Diff':>10} | {'Dist_Diff':>10} | {'PG_Diff':>8} | {'GateNum B':>9} | {'GateNum D':>9}")
print("-" * 100)

# 排除无差别电路(如 linear / star 小规模)
interesting = [x for x in diff_data if abs(x['fid_diff']) > 1e-8]
for dd in interesting:
    print(f"{dd['circ']:>12} | {dd['fid_diff']:>+10.6f} | {dd['deco_diff']:>+10.6f} | {dd['cz_diff']:>+10.6f} | {dd['idle_diff']:>+10.2f} | {dd['dist_diff']:>+10.2f} | {dd['pg_diff']:>+8} | {dd['base_gate_num'] or 0:>9} | {dd['dual_gate_num'] or 0:>9}")

# ======= 关键分析: 分解 F_final = F_deco * F_cz (忽略F_trans=1) =======
print("\n\n" + "=" * 100)
print("分解分析: 哪个保真度组分主导了差异?")
print("=" * 100)
print(f"{'电路':>12} | {'ΔF_final':>10} | {'ΔF来自退相干':>14} | {'ΔF来自CZ门':>12} | {'主导因素':>10}")
print("-" * 100)

for dd in interesting:
    b_fid = dd['base_gate_num'] and dd['base_idle'] is not None
    d_fid = dd['dual_gate_num'] and dd['dual_idle'] is not None
    
    if not (b_fid and d_fid):
        continue
    
    # F = F_deco * F_cz
    # ΔF ≈ ΔF_deco * F_cz + F_deco * ΔF_cz (一阶近似)
    # 更精确: ΔF = F_dual - F_base = F_deco_dual * F_cz_dual - F_deco_base * F_cz_base
    
    f_deco_b = math.exp(-dd['base_idle'] / T_eff) if dd['base_idle'] else 1.0
    f_deco_d = math.exp(-dd['dual_idle'] / T_eff) if dd['dual_idle'] else 1.0
    f_cz_b = F_cz_val ** dd['base_gate_num'] if dd['base_gate_num'] else 1.0
    f_cz_d = F_cz_val ** dd['dual_gate_num'] if dd['dual_gate_num'] else 1.0
    
    f_base = f_deco_b * f_cz_b
    f_dual = f_deco_d * f_cz_d
    delta_f = f_dual - f_base
    
    # 分解为退相干贡献和CZ门贡献
    # ΔF = (F_deco_d - F_deco_b) * F_cz_avg + F_deco_avg * (F_cz_d - F_cz_b)
    f_cz_avg = (f_cz_b + f_cz_d) / 2
    f_deco_avg = (f_deco_b + f_deco_d) / 2
    
    delta_from_deco = (f_deco_d - f_deco_b) * f_cz_avg
    delta_from_cz = f_deco_avg * (f_cz_d - f_cz_b)
    
    if abs(delta_from_deco) >= abs(delta_from_cz):
        dominant = "退相干"
    else:
        dominant = "CZ门"
    
    print(f"{dd['circ']:>12} | {delta_f:>+10.6f} | {delta_from_deco:>+14.6f} | {delta_from_cz:>+12.6f} | {dominant:>10}")

# ======= 进一步: 退相干差异的来源分解 =======
print("\n\n" + "=" * 100)
print("退相干差异来源分解: t_idle = num_q * t_total - gate_num * T_cz")
print("t_total = PG_steps * T_cz + move_steps * 4 * T_trans + sum(max_dis/Move_speed)")
print("=" * 100)
print(f"{'电路':>12} | {'Base PG':>8} | {'Dual PG':>8} | {'ΔPG':>5} | {'Base MoveS':>10} | {'Dual MoveS':>10} | {'ΔMoveS':>7} | {'Base Dist':>10} | {'Dual Dist':>10} | {'ΔDist':>10}")
print("-" * 100)

for dd in interesting:
    print(f"{dd['circ']:>12} | {dd['base_pg'] or 0:>8} | {dd['dual_pg'] or 0:>8} | {dd['pg_diff']:>+5} | {dd['base_move_steps'] or 0:>10} | {dd['dual_move_steps'] or 0:>10} | {(dd['dual_move_steps'] or 0) - (dd['base_move_steps'] or 0):>+7} | {dd['base_dist'] or 0:>10.2f} | {dd['dual_dist'] or 0:>10.2f} | {dd['dist_diff']:>+10.2f}")

print("\n\n[结论分析]")
print("由于 F_trans = 1 (无传输损失), CZ门数量两边相同 (相同电路)")
print("保真度差异完全来自退相干 F_deco = exp(-t_idle/T_eff)")
print("t_idle 差异又由以下因素决定:")
print("  1. parallel_gates 分组数 (PG steps) — 更多分组 => t_total更大 => t_idle更大")
print("  2. 移动步数 (movement steps) — 更多移动 => t_total更大")  
print("  3. 移动距离 (move distance) — 更远移动 => t_total更大")
print("  以上三者的根源都是: embedding质量 (初始映射+placement)")
